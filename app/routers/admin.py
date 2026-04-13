from fastapi import APIRouter, Depends, HTTPException, Cookie, Request, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from datetime import datetime, date, timedelta
import os, cloudinary, cloudinary.uploader
from app.database import get_db
from app.core.config import TZ
from app.routers.auth import verificar_sesion, obtener_rol
from app.models.pedidos import Pedido, ItemPedido
from app.models.productos import Producto, Categoria, ProductoVariante
import logging
logger = logging.getLogger("floreria")
from app.models.clientes import Cliente
from app.models.configuracion import CodigoDescuento
from app.models.usuarios import Usuario
from app.models.egresos import Egreso, GastoRecurrente, MetodoPagoEgreso, OtroIngreso, CategoriaGasto
from app.models.banners import BannerCatalogo
from app.models.cuentas import CuentaTransferencia, CuentaFinanciera, MovimientoCuenta
from app.models.fiscales import DatosFiscalesCliente
from app.models.proveedores import Proveedor

router = APIRouter()

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME", "ddku2wmpk"),
    api_key=os.getenv("CLOUDINARY_API_KEY", "543563876228939"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET", ""),
)


def _auth(panel_session):
    if not verificar_sesion(panel_session):
        raise HTTPException(status_code=401, detail="No autenticado")
    rol = obtener_rol(panel_session)
    if rol != "admin":
        raise HTTPException(status_code=403, detail="Acceso denegado: se requiere rol admin")


# ══════ USUARIOS ══════

@router.get("/usuarios")
async def listar_usuarios(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(text("SELECT id, nombre, username, rol, activo, created_at FROM usuarios ORDER BY id"))
    return [
        {"id": r[0], "nombre": r[1], "username": r[2], "rol": r[3],
         "activo": r[4], "created_at": str(r[5]) if r[5] else None}
        for r in result.fetchall()
    ]


@router.post("/usuarios")
async def crear_usuario(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    import hashlib
    password = data.get("password", "")
    if not password:
        raise HTTPException(status_code=400, detail="Contraseña requerida")
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    try:
        await db.execute(text(
            "INSERT INTO usuarios (nombre, username, password_hash, rol, activo) VALUES (:n, :u, :p, :r, true)"
        ), {"n": data["nombre"], "u": data["username"], "p": pw_hash, "r": data.get("rol", "operador")})
        await db.commit()
    except Exception as e:
        await db.rollback()
        if "unique" in str(e).lower() or "duplicate" in str(e).lower():
            raise HTTPException(status_code=400, detail="Username ya existe")
        raise HTTPException(status_code=500, detail=str(e))
    return {"ok": True}


@router.put("/usuarios/{user_id}")
async def actualizar_usuario(
    user_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    # Prevent deactivating last admin
    if "activo" in data and not data["activo"]:
        r = await db.execute(text("SELECT rol FROM usuarios WHERE id = :id"), {"id": user_id})
        row = r.fetchone()
        if row and row[0] == "admin":
            cnt = await db.execute(text("SELECT COUNT(*) FROM usuarios WHERE rol='admin' AND activo=true"))
            if (cnt.scalar() or 0) <= 1:
                raise HTTPException(status_code=400, detail="No se puede desactivar el último admin")
    sets = []
    params = {"id": user_id}
    for k in ["nombre", "rol", "activo"]:
        if k in data:
            sets.append(f"{k} = :{k}")
            params[k] = data[k]
    if not sets: return {"ok": True}
    await db.execute(text(f"UPDATE usuarios SET {', '.join(sets)} WHERE id = :id"), params)
    await db.commit()
    return {"ok": True}


@router.post("/usuarios/{user_id}/cambiar-password")
async def cambiar_password(
    user_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    password = data.get("password", "")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Contraseña muy corta (mínimo 8 caracteres)")
    import hashlib
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    await db.execute(text("UPDATE usuarios SET password_hash = :p WHERE id = :id"), {"p": pw_hash, "id": user_id})
    await db.commit()
    return {"ok": True}


# ══════ EGRESOS ══════

@router.post("/egresos/importar-kyte")
async def importar_egresos_kyte(
    file: UploadFile = File(...),
    dry_run: bool = True,
    desde: str | None = None,
    hasta: str | None = None,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """Importa egresos desde un export xlsx de Kyte.
    Columnas esperadas: ID, Pagado, Fecha/Hora (creación), Nombre del gasto,
    Valor, Categoría, Proveedor, Fecha de vencimiento, Fecha de pago,
    Fecha de referencia, Recurrente, Observación.
    Dedupe por ID Kyte guardado en columna 'referencia'.
    Si dry_run=true, devuelve preview sin escribir nada.
    Filtra por rango opcional [desde, hasta]."""
    _auth(panel_session)
    import openpyxl
    from io import BytesIO
    contenido = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer xlsx: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    headers = [str(h or '').strip().lower() for h in rows[0]]

    def col(name_substring):
        for i, h in enumerate(headers):
            if name_substring in h:
                return i
        return -1

    idx_id = col("id")
    idx_pagado = col("pagado")
    idx_nombre = col("nombre")
    idx_valor = col("valor")
    idx_cat = col("categor")
    idx_prov = col("proveedor")
    idx_fpago = col("fecha de pago")
    idx_recurrente = col("recurrente")
    idx_obs = col("observa")

    if min(idx_id, idx_nombre, idx_valor, idx_fpago) < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas faltantes. Headers detectados: {headers}",
        )

    # Mapeo categorías Kyte → categorías sistema
    cat_map = {
        "flor": "Flor",
        "funcion": "Nomina",  # Funcionários
        "transporte": "Transporte",
        "log": "Transporte",
        "gastos del hogar": "Otros",
        "insumos": "Insumos",
        "servicios": "Servicios",
        "mantenimiento": "Mantenimiento",
        "impuestos": "Otros",
        "otros": "Otros",
    }

    def map_cat(raw):
        if not raw:
            return "Otros"
        r = str(raw).strip().lower()
        for k, v in cat_map.items():
            if k in r:
                return v
        return "Otros"

    def map_metodo_pago(obs):
        if not obs:
            return None
        o = str(obs).strip().lower()
        if o in ("cc", "caja chica"):
            return "Caja chica"
        if o in ("efectivo", "caja"):
            return "Caja"
        if "tdc" in o:
            return None  # ambiguo, Fer edita después
        if "transfer" in o:
            return None
        return None

    def parse_valor(v):
        """Kyte usa formato europeo: coma=decimal, punto=miles.
           "30.000" -> 30000  |  "713,03" -> 713.03  |  "1.234,56" -> 1234.56
           "850" -> 850  |  numeros nativos int/float pasan tal cual."""
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return int(round(float(v) * 100))
        s = str(v).strip().replace(" ", "").replace("$", "")
        if not s:
            return 0
        if "," in s:
            # Coma = decimal. Puntos = miles, los quitamos.
            s = s.replace(".", "").replace(",", ".")
        else:
            # Sin coma: puntos son separador de miles (formato Kyte). Quitar.
            s = s.replace(".", "")
        try:
            return int(round(float(s) * 100))
        except Exception:
            return 0

    def parse_fecha(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v).strip()
        # Formatos posibles: "1/4/2025", "01/04/2025", "2025-04-01"
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    desde_d = date.fromisoformat(desde) if desde else None
    hasta_d = date.fromisoformat(hasta) if hasta else None

    # Procesar filas
    a_importar = []
    skipped_filtro = 0
    skipped_invalido = 0
    for row in rows[1:]:
        if not row or row[idx_id] is None:
            continue
        kyte_id = str(row[idx_id]).strip()
        fecha = parse_fecha(row[idx_fpago])
        if not fecha:
            skipped_invalido += 1
            continue
        if desde_d and fecha < desde_d:
            skipped_filtro += 1
            continue
        if hasta_d and fecha > hasta_d:
            skipped_filtro += 1
            continue
        monto = parse_valor(row[idx_valor])
        if monto <= 0:
            skipped_invalido += 1
            continue
        concepto = str(row[idx_nombre] or "").strip() or "(sin nombre)"
        cat_raw = row[idx_cat] if idx_cat >= 0 else None
        proveedor = str(row[idx_prov] or "").strip() if idx_prov >= 0 else None
        proveedor = proveedor or None
        recurrente = False
        if idx_recurrente >= 0:
            recurrente = str(row[idx_recurrente] or "").strip().lower() in ("si", "sí", "yes", "true", "1")
        obs = str(row[idx_obs] or "").strip() if idx_obs >= 0 else None
        obs = obs or None
        a_importar.append({
            "kyte_id": kyte_id,
            "fecha": fecha,
            "concepto": concepto,
            "categoria": map_cat(cat_raw),
            "categoria_original": str(cat_raw or ""),
            "monto": monto,
            "metodo_pago": map_metodo_pago(obs),
            "proveedor": proveedor,
            "notas": obs,
            "es_recurrente": recurrente,
        })

    # Dedupe contra BD existente por referencia
    existing_set = set()
    if a_importar:
        kyte_ids = [r["kyte_id"] for r in a_importar]
        existing = (await db.execute(text(
            "SELECT referencia FROM egresos WHERE referencia = ANY(:ids)"
        ), {"ids": kyte_ids})).fetchall()
        existing_set = {row[0] for row in existing}
    nuevos = [r for r in a_importar if r["kyte_id"] not in existing_set]
    a_actualizar = [r for r in a_importar if r["kyte_id"] in existing_set]

    # Resumen
    total_monto = sum(r["monto"] for r in a_importar)
    por_categoria = {}
    for r in a_importar:
        por_categoria[r["categoria"]] = por_categoria.get(r["categoria"], 0) + r["monto"]
    por_metodo = {}
    for r in a_importar:
        m = r["metodo_pago"] or "(sin método)"
        por_metodo[m] = por_metodo.get(m, 0) + r["monto"]

    resumen = {
        "total_filas_xlsx": len(rows) - 1,
        "skipped_filtro_fecha": skipped_filtro,
        "skipped_invalido": skipped_invalido,
        "a_insertar_nuevos": len(nuevos),
        "a_actualizar_existentes": len(a_actualizar),
        "a_importar": len(a_importar),
        "monto_total": total_monto,
        "por_categoria": por_categoria,
        "por_metodo_pago": por_metodo,
        "primeros_5": [
            {"fecha": str(r["fecha"]), "concepto": r["concepto"],
             "categoria": r["categoria"], "monto": r["monto"],
             "metodo_pago": r["metodo_pago"], "proveedor": r["proveedor"]}
            for r in a_importar[:5]
        ],
        "dry_run": dry_run,
    }

    if dry_run or not a_importar:
        return resumen

    # UPSERT: insert nuevos, update existentes
    insertados = 0
    actualizados = 0
    try:
        for r in nuevos:
            await db.execute(text("""
                INSERT INTO egresos
                    (fecha, concepto, categoria, monto, metodo_pago, proveedor, notas, referencia, es_recurrente)
                VALUES (:f, :c, :cat, :m, :mp, :prov, :n, :ref, :er)
            """), {
                "f": r["fecha"], "c": r["concepto"], "cat": r["categoria"],
                "m": r["monto"], "mp": r["metodo_pago"],
                "prov": r["proveedor"], "n": r["notas"],
                "ref": r["kyte_id"], "er": r["es_recurrente"],
            })
            insertados += 1
        for r in a_actualizar:
            await db.execute(text("""
                UPDATE egresos SET
                    fecha = :f, concepto = :c, categoria = :cat,
                    monto = :m, metodo_pago = :mp, proveedor = :prov,
                    notas = :n, es_recurrente = :er
                WHERE referencia = :ref
            """), {
                "f": r["fecha"], "c": r["concepto"], "cat": r["categoria"],
                "m": r["monto"], "mp": r["metodo_pago"],
                "prov": r["proveedor"], "n": r["notas"],
                "ref": r["kyte_id"], "er": r["es_recurrente"],
            })
            actualizados += 1
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Importación falló (insertados={insertados}, actualizados={actualizados}): {e}"
        )

    resumen["insertados"] = insertados
    resumen["actualizados"] = actualizados
    return resumen


@router.get("/egresos")
async def listar_egresos(
    desde: str | None = None, hasta: str | None = None,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    sql = "SELECT id, fecha, concepto, categoria, monto, metodo_pago, proveedor, notas, referencia, es_recurrente, cuenta_id FROM egresos"
    params = {}
    wheres = []
    if desde:
        wheres.append("fecha >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        wheres.append("fecha <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)
    if wheres:
        sql += " WHERE " + " AND ".join(wheres)
    sql += " ORDER BY fecha DESC LIMIT 500"
    result = await db.execute(text(sql), params)
    return [
        {"id": r[0], "fecha": str(r[1]), "concepto": r[2], "categoria": r[3],
         "monto": r[4], "metodo_pago": r[5], "proveedor": r[6], "notas": r[7],
         "referencia": r[8], "es_recurrente": r[9], "cuenta_id": r[10]}
        for r in result.fetchall()
    ]


@router.post("/egresos")
async def crear_egreso(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    try:
        fecha_val = date.fromisoformat(data["fecha"]) if isinstance(data["fecha"], str) else data["fecha"]
        await db.execute(text("""
            INSERT INTO egresos (fecha, concepto, categoria, monto, metodo_pago, proveedor, notas, referencia, es_recurrente, cuenta_id)
            VALUES (:f, :c, :cat, :m, :mp, :prov, :n, :ref, :er, :ci)
        """), {
            "f": fecha_val, "c": data["concepto"], "cat": data.get("categoria", "otro"),
            "m": data.get("monto", 0), "mp": data.get("metodo_pago"),
            "prov": data.get("proveedor"), "n": data.get("notas"),
            "ref": data.get("referencia"), "er": data.get("es_recurrente", False),
            "ci": data.get("cuenta_id"),
        })
        await db.commit()
        return {"ok": True}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/egresos/{egreso_id}")
async def actualizar_egreso(
    egreso_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    sets = []
    params = {"id": egreso_id}
    for k in ["concepto", "categoria", "monto", "metodo_pago", "proveedor", "notas", "referencia", "cuenta_id"]:
        if k in data:
            sets.append(f"{k} = :{k}")
            params[k] = data[k]
    if "fecha" in data:
        sets.append("fecha = :fecha")
        params["fecha"] = date.fromisoformat(data["fecha"]) if isinstance(data["fecha"], str) else data["fecha"]
    if not sets:
        return {"ok": True}
    await db.execute(text(f"UPDATE egresos SET {', '.join(sets)} WHERE id = :id"), params)
    await db.commit()
    return {"ok": True}


@router.delete("/egresos/{egreso_id}")
async def eliminar_egreso(
    egreso_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    await db.execute(text("DELETE FROM egresos WHERE id = :id"), {"id": egreso_id})
    await db.commit()
    return {"ok": True}


# --- Gastos recurrentes ---

@router.get("/gastos-recurrentes")
async def listar_gastos_recurrentes(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(GastoRecurrente).order_by(GastoRecurrente.nombre))
    return [
        {"id": g.id, "nombre": g.nombre, "categoria": g.categoria,
         "frecuencia": g.frecuencia, "monto_sugerido": g.monto_sugerido,
         "metodo_pago": g.metodo_pago, "proveedor": g.proveedor, "activo": g.activo}
        for g in result.scalars().all()
    ]


@router.post("/gastos-recurrentes")
async def crear_gasto_recurrente(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    g = GastoRecurrente(
        nombre=data["nombre"], categoria=data.get("categoria", "otro"),
        frecuencia=data.get("frecuencia", "mensual"),
        monto_sugerido=data.get("monto_sugerido", 0),
        metodo_pago=data.get("metodo_pago") or None,
        proveedor=data.get("proveedor") or None, activo=True,
    )
    db.add(g)
    await db.commit()
    await db.refresh(g)
    return {"ok": True, "id": g.id}


@router.put("/gastos-recurrentes/{gasto_id}")
async def actualizar_gasto_recurrente(
    gasto_id: int, request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(GastoRecurrente).where(GastoRecurrente.id == gasto_id))
    g = result.scalar_one_or_none()
    if not g:
        raise HTTPException(status_code=404, detail="No encontrado")
    data = await request.json()
    for k in ["nombre", "categoria", "frecuencia", "monto_sugerido", "metodo_pago", "proveedor", "activo"]:
        if k in data:
            setattr(g, k, data[k])
    await db.commit()
    return {"ok": True}


@router.delete("/gastos-recurrentes/{gasto_id}")
async def eliminar_gasto_recurrente(
    gasto_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(GastoRecurrente).where(GastoRecurrente.id == gasto_id))
    g = result.scalar_one_or_none()
    if not g:
        raise HTTPException(status_code=404, detail="No encontrado")
    await db.delete(g)
    await db.commit()
    return {"ok": True}


# --- Proveedores ---

@router.get("/proveedores")
async def listar_proveedores(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(Proveedor).order_by(Proveedor.nombre))
    return [{"id":p.id,"nombre":p.nombre,"contacto":p.contacto,"telefono":p.telefono,"notas":p.notas,"activo":p.activo} for p in result.scalars().all()]

@router.post("/proveedores")
async def crear_proveedor(request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    data = await request.json()
    p = Proveedor(nombre=data["nombre"].strip(), contacto=data.get("contacto"), telefono=data.get("telefono"), notas=data.get("notas"))
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return {"ok":True,"id":p.id}

@router.put("/proveedores/{prov_id}")
async def actualizar_proveedor(prov_id: int, request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(Proveedor).where(Proveedor.id == prov_id))
    p = result.scalar_one_or_none()
    if not p: raise HTTPException(status_code=404, detail="No encontrado")
    data = await request.json()
    for k in ["nombre","contacto","telefono","notas","activo"]:
        if k in data: setattr(p, k, data[k])
    await db.commit()
    return {"ok":True}

@router.delete("/proveedores/{prov_id}")
async def eliminar_proveedor(prov_id: int, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    # Check if has egresos
    cnt = await db.execute(text("SELECT COUNT(*) FROM egresos WHERE proveedor = (SELECT nombre FROM proveedores WHERE id = :id)"), {"id": prov_id})
    if (cnt.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="Tiene egresos asociados — desactívalo en vez de eliminar")
    result = await db.execute(select(Proveedor).where(Proveedor.id == prov_id))
    p = result.scalar_one_or_none()
    if not p: raise HTTPException(status_code=404, detail="No encontrado")
    await db.delete(p)
    await db.commit()
    return {"ok":True}

# --- Categorías de gasto ---

@router.get("/categorias-gasto")
async def listar_categorias_gasto(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(CategoriaGasto).order_by(CategoriaGasto.nombre))
    cats = result.scalars().all()
    items = []
    for c in cats:
        cnt = await db.execute(text("SELECT COUNT(*) FROM egresos WHERE categoria = :n"), {"n": c.nombre})
        items.append({"id": c.id, "nombre": c.nombre, "activo": c.activo, "egresos": cnt.scalar() or 0})
    return items

@router.post("/categorias-gasto")
async def crear_categoria_gasto(request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    data = await request.json()
    c = CategoriaGasto(nombre=data["nombre"].strip(), activo=True)
    db.add(c)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=400, detail="Ya existe")
    return {"ok": True, "id": c.id}

@router.put("/categorias-gasto/{cat_id}")
async def actualizar_categoria_gasto(cat_id: int, request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(CategoriaGasto).where(CategoriaGasto.id == cat_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="No encontrada")
    data = await request.json()
    old_name = c.nombre
    if "nombre" in data:
        c.nombre = data["nombre"].strip()
        # Update egresos with old name
        if c.nombre != old_name:
            await db.execute(text("UPDATE egresos SET categoria = :new WHERE categoria = :old"), {"new": c.nombre, "old": old_name})
    if "activo" in data:
        c.activo = data["activo"]
    await db.commit()
    return {"ok": True}

@router.delete("/categorias-gasto/{cat_id}")
async def eliminar_categoria_gasto(cat_id: int, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(CategoriaGasto).where(CategoriaGasto.id == cat_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="No encontrada")
    cnt = await db.execute(text("SELECT COUNT(*) FROM egresos WHERE categoria = :n"), {"n": c.nombre})
    if (cnt.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="Tiene egresos asociados")
    await db.delete(c)
    await db.commit()
    return {"ok": True}

# --- Métodos de pago egresos ---

@router.get("/metodos-pago-egreso")
async def listar_metodos_pago_egreso(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(MetodoPagoEgreso).order_by(MetodoPagoEgreso.nombre))
    return [{"id": m.id, "nombre": m.nombre, "activo": m.activo} for m in result.scalars().all()]


@router.post("/metodos-pago-egreso")
async def crear_metodo_pago_egreso(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    m = MetodoPagoEgreso(nombre=data["nombre"], activo=True)
    db.add(m)
    await db.commit()
    await db.refresh(m)
    return {"ok": True, "id": m.id}


@router.put("/metodos-pago-egreso/{mp_id}")
async def actualizar_metodo_pago_egreso(
    mp_id: int, request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(MetodoPagoEgreso).where(MetodoPagoEgreso.id == mp_id))
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="No encontrado")
    data = await request.json()
    if "nombre" in data: m.nombre = data["nombre"]
    if "activo" in data: m.activo = data["activo"]
    await db.commit()
    return {"ok": True}


@router.delete("/metodos-pago-egreso/{mp_id}")
async def eliminar_metodo_pago_egreso(
    mp_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(MetodoPagoEgreso).where(MetodoPagoEgreso.id == mp_id))
    m = result.scalar_one_or_none()
    if not m:
        raise HTTPException(status_code=404, detail="No encontrado")
    # Check if has associated egresos
    count = await db.execute(select(func.count(Egreso.id)).where(Egreso.metodo_pago == m.nombre))
    if (count.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="Tiene egresos registrados — desactívalo en vez de eliminar")
    await db.delete(m)
    await db.commit()
    return {"ok": True}


# --- Exportar egresos Excel ---

@router.get("/egresos/exportar")
async def exportar_egresos(
    desde: str | None = None, hasta: str | None = None,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    from fastapi.responses import StreamingResponse
    import io, openpyxl
    query = select(Egreso).order_by(Egreso.fecha.desc())
    if desde: query = query.where(Egreso.fecha >= desde)
    if hasta: query = query.where(Egreso.fecha <= hasta)
    result = await db.execute(query)
    egresos = result.scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Egresos"
    headers = ["Fecha", "Concepto", "Categoría", "Método de pago", "Monto", "Notas"]
    ws.append(headers)
    for c in range(1, len(headers)+1): ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for e in egresos:
        ws.append([str(e.fecha), e.concepto, e.categoria, e.metodo_pago or "", round(e.monto/100, 2), e.notas or ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    hoy = datetime.now(TZ).strftime("%Y-%m-%d")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=egresos_{hoy}.xlsx"})


# --- Otros ingresos ---

@router.get("/otros-ingresos")
async def listar_otros_ingresos(
    desde: str | None = None, hasta: str | None = None,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    query = select(OtroIngreso).order_by(OtroIngreso.fecha.desc())
    if desde: query = query.where(OtroIngreso.fecha >= desde)
    if hasta: query = query.where(OtroIngreso.fecha <= hasta)
    result = await db.execute(query)
    return [
        {"id": o.id, "fecha": str(o.fecha), "concepto": o.concepto,
         "monto": o.monto, "metodo_pago": o.metodo_pago, "notas": o.notas}
        for o in result.scalars().all()
    ]


@router.post("/otros-ingresos")
async def crear_otro_ingreso(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    o = OtroIngreso(
        fecha=data["fecha"], concepto=data["concepto"],
        monto=data.get("monto", 0), metodo_pago=data.get("metodo_pago"),
        notas=data.get("notas"),
    )
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return {"ok": True, "id": o.id}


@router.delete("/otros-ingresos/{oi_id}")
async def eliminar_otro_ingreso(
    oi_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(OtroIngreso).where(OtroIngreso.id == oi_id))
    o = result.scalar_one_or_none()
    if not o:
        raise HTTPException(status_code=404, detail="No encontrado")
    await db.delete(o)
    await db.commit()
    return {"ok": True}


# --- Flujo de caja ---

@router.get("/finanzas/flujo-caja")
async def flujo_caja(
    desde: str, hasta: str,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    # Ingresos por día (fecha que entró el dinero, no fecha entrega)
    ingresos = await db.execute(text("""
        SELECT pago_confirmado_at::date as fecha, COALESCE(SUM(total),0) as total
        FROM pedidos WHERE pago_confirmado_at::date BETWEEN :d AND :h
          AND estado NOT IN ('Cancelado','rechazado') AND pago_confirmado = true
        GROUP BY pago_confirmado_at::date ORDER BY fecha
    """), _dp(desde, hasta))
    ing_map = {str(r[0]): r[1] for r in ingresos.fetchall()}
    # Otros ingresos por día
    otros = await db.execute(text("""
        SELECT fecha::date as fecha, COALESCE(SUM(monto),0) as total
        FROM otros_ingresos WHERE fecha BETWEEN :d AND :h
        GROUP BY fecha::date ORDER BY fecha
    """), _dp(desde, hasta))
    for r in otros.fetchall():
        ing_map[str(r[0])] = ing_map.get(str(r[0]), 0) + r[1]
    # Egresos por día
    egresos = await db.execute(text("""
        SELECT fecha::date as fecha, COALESCE(SUM(monto),0) as total
        FROM egresos WHERE fecha BETWEEN :d AND :h
        GROUP BY fecha::date ORDER BY fecha
    """), _dp(desde, hasta))
    egr_map = {str(r[0]): r[1] for r in egresos.fetchall()}
    # Build day-by-day
    d = date.fromisoformat(desde)
    h = date.fromisoformat(hasta)
    days = []
    acum = 0
    while d <= h:
        ds = str(d)
        ing = ing_map.get(ds, 0)
        egr = egr_map.get(ds, 0)
        saldo = ing - egr
        acum += saldo
        days.append({"fecha": ds, "ingresos": ing, "egresos": egr, "saldo": saldo, "acumulado": acum})
        d += timedelta(days=1)
    return days


# ══════ DESCUENTOS ══════

@router.get("/descuentos")
async def listar_descuentos(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(CodigoDescuento).order_by(CodigoDescuento.id.desc()))
    return [
        {"id": d.id, "codigo": d.codigo, "tipo": d.tipo, "valor": d.valor,
         "descripcion": d.descripcion, "activo": d.activo,
         "fecha_inicio": str(d.fecha_inicio) if d.fecha_inicio else None,
         "fecha_expiracion": str(d.fecha_expiracion) if d.fecha_expiracion else None,
         "usos_maximos": d.usos_maximos, "usos_actuales": d.usos_actuales}
        for d in result.scalars().all()
    ]


@router.post("/descuentos")
async def crear_descuento(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    d = CodigoDescuento(
        codigo=data["codigo"].strip().upper(),
        tipo=data.get("tipo", "porcentaje"),
        valor=data.get("valor", 0),
        descripcion=data.get("descripcion"),
        activo=data.get("activo", True),
        fecha_inicio=data.get("fecha_inicio"),
        fecha_expiracion=data.get("fecha_expiracion"),
        usos_maximos=data.get("usos_maximos"),
    )
    db.add(d)
    await db.commit()
    await db.refresh(d)
    return {"ok": True, "id": d.id}


@router.put("/descuentos/{desc_id}")
async def actualizar_descuento(
    desc_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(CodigoDescuento).where(CodigoDescuento.id == desc_id))
    d = result.scalar_one_or_none()
    if not d:
        raise HTTPException(status_code=404, detail="Descuento no encontrado")
    data = await request.json()
    for k in ["codigo", "tipo", "valor", "descripcion", "activo", "fecha_inicio", "fecha_expiracion", "usos_maximos"]:
        if k in data:
            setattr(d, k, data[k])
    await db.commit()
    return {"ok": True}


@router.delete("/descuentos/{desc_id}")
async def eliminar_descuento(
    desc_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(CodigoDescuento).where(CodigoDescuento.id == desc_id))
    d = result.scalar_one_or_none()
    if not d:
        raise HTTPException(status_code=404, detail="Descuento no encontrado")
    await db.delete(d)
    await db.commit()
    return {"ok": True}


# ══════ BANNERS ══════

@router.get("/banners")
async def listar_banners(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(BannerCatalogo).order_by(BannerCatalogo.orden))
    return [
        {"id": b.id, "imagen_url": b.imagen_url, "titulo": b.titulo,
         "subtitulo": b.subtitulo, "link": b.link, "orden": b.orden, "activo": b.activo}
        for b in result.scalars().all()
    ]


@router.post("/banners")
async def crear_banner(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    b = BannerCatalogo(
        imagen_url=data["imagen_url"],
        titulo=data.get("titulo"),
        subtitulo=data.get("subtitulo"),
        link=data.get("link"),
        orden=data.get("orden", 0),
        activo=data.get("activo", True),
    )
    db.add(b)
    await db.commit()
    await db.refresh(b)
    return {"ok": True, "id": b.id}


@router.delete("/banners/{banner_id}")
async def eliminar_banner(
    banner_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(BannerCatalogo).where(BannerCatalogo.id == banner_id))
    b = result.scalar_one_or_none()
    if not b:
        raise HTTPException(status_code=404, detail="Banner no encontrado")
    await db.delete(b)
    await db.commit()
    return {"ok": True}


# ══════ FACTURACIÓN ══════

def _fact_row(r):
    return {"id":r[0],"folio":r[1],"fecha":str(r[2]) if r[2] else None,"cliente":r[3],
     "canal":r[4],"subtotal":r[5],"iva":round(r[5]*0.16),"total":r[5]+round(r[5]*0.16),
     "estado":r[6],"datos_fiscales_id":r[7],"folio_fiscal":r[8],"customer_id":r[9]}

@router.get("/facturacion/pendientes")
async def facturacion_pendientes(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(text("""
        SELECT p.id, p.numero, p.fecha_entrega, COALESCE(c.nombre,'Mostrador'),
               p.canal, p.total, p.estado, p.datos_fiscales_id, p.folio_fiscal, p.customer_id
        FROM pedidos p LEFT JOIN clientes c ON c.id=p.customer_id
        WHERE p.requiere_factura = true AND (p.facturado = false OR p.facturado IS NULL)
        ORDER BY p.fecha_entrega DESC
    """))
    return [_fact_row(r) for r in result.fetchall()]

@router.get("/facturacion/facturados")
async def facturacion_facturados(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(text("""
        SELECT p.id, p.numero, p.fecha_entrega, COALESCE(c.nombre,'Mostrador'),
               p.canal, p.total, p.estado, p.datos_fiscales_id, p.folio_fiscal, p.customer_id
        FROM pedidos p LEFT JOIN clientes c ON c.id=p.customer_id
        WHERE p.facturado = true ORDER BY p.fecha_entrega DESC
    """))
    return [_fact_row(r) for r in result.fetchall()]

@router.get("/facturacion/count")
async def facturacion_count(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(text("SELECT COUNT(*) FROM pedidos WHERE requiere_factura = true AND (facturado = false OR facturado IS NULL)"))
    return {"count": result.scalar() or 0}

@router.post("/facturacion/{pedido_id}/marcar")
async def marcar_facturado(pedido_id: int, request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    data = await request.json()
    folio_fiscal = data.get("folio_fiscal", "")
    await db.execute(text("UPDATE pedidos SET facturado = true, folio_fiscal = :ff WHERE id = :id"), {"id": pedido_id, "ff": folio_fiscal or None})
    await db.commit()
    return {"ok": True}

# --- Datos fiscales ---

@router.get("/datos-fiscales/{cliente_id}")
async def obtener_datos_fiscales(cliente_id: int, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(DatosFiscalesCliente).where(DatosFiscalesCliente.cliente_id == cliente_id))
    df = result.scalar_one_or_none()
    if not df: return {"existe": False}
    return {"existe":True,"id":df.id,"rfc":df.rfc,"razon_social":df.razon_social,"regimen_fiscal":df.regimen_fiscal,
            "uso_cfdi":df.uso_cfdi,"correo_fiscal":df.correo_fiscal,"codigo_postal":df.codigo_postal}

@router.post("/datos-fiscales")
async def guardar_datos_fiscales(request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    data = await request.json()
    cliente_id = data.get("cliente_id")
    # Upsert
    result = await db.execute(select(DatosFiscalesCliente).where(DatosFiscalesCliente.cliente_id == cliente_id))
    df = result.scalar_one_or_none()
    if not df:
        df = DatosFiscalesCliente(cliente_id=cliente_id)
        db.add(df)
    for k in ["rfc","razon_social","regimen_fiscal","uso_cfdi","correo_fiscal","codigo_postal"]:
        if k in data: setattr(df, k, data[k])
    await db.commit()
    await db.refresh(df)
    return {"ok":True,"id":df.id}

@router.get("/datos-fiscales/pedido/{pedido_id}")
async def datos_fiscales_pedido(pedido_id: int, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    # Get datos_fiscales_id from pedido
    r = await db.execute(text("SELECT datos_fiscales_id FROM pedidos WHERE id = :id"), {"id": pedido_id})
    row = r.fetchone()
    if not row or not row[0]: return {"existe": False}
    result = await db.execute(select(DatosFiscalesCliente).where(DatosFiscalesCliente.id == row[0]))
    df = result.scalar_one_or_none()
    if not df: return {"existe": False}
    return {"existe":True,"rfc":df.rfc,"razon_social":df.razon_social,"regimen_fiscal":df.regimen_fiscal,
            "uso_cfdi":df.uso_cfdi,"correo_fiscal":df.correo_fiscal,"codigo_postal":df.codigo_postal}

@router.get("/catalogos-fiscales")
async def catalogos_fiscales(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    regs = await db.execute(text("SELECT codigo, nombre FROM regimenes_fiscales ORDER BY codigo"))
    usos = await db.execute(text("SELECT codigo, nombre FROM usos_cfdi ORDER BY codigo"))
    return {"regimenes": [{"codigo":r[0],"nombre":r[1]} for r in regs.fetchall()],
            "usos": [{"codigo":r[0],"nombre":r[1]} for r in usos.fetchall()]}


# ══════ CUENTAS TRANSFERENCIA ══════

@router.get("/cuentas-transferencia")
async def listar_cuentas(panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(CuentaTransferencia).order_by(CuentaTransferencia.id))
    return [{"id":c.id,"banco":c.banco,"titular":c.titular,"tarjeta":c.tarjeta,"clabe":c.clabe,"activa":c.activa} for c in result.scalars().all()]

@router.post("/cuentas-transferencia")
async def crear_cuenta(request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    data = await request.json()
    c = CuentaTransferencia(banco=data["banco"], titular=data.get("titular",""), tarjeta=data.get("tarjeta",""), clabe=data.get("clabe",""), activa=False)
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return {"ok":True,"id":c.id}

@router.put("/cuentas-transferencia/{cuenta_id}")
async def actualizar_cuenta(cuenta_id: int, request: Request, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(CuentaTransferencia).where(CuentaTransferencia.id == cuenta_id))
    c = result.scalar_one_or_none()
    if not c: raise HTTPException(status_code=404, detail="No encontrada")
    data = await request.json()
    for k in ["banco","titular","tarjeta","clabe"]:
        if k in data: setattr(c, k, data[k])
    await db.commit()
    return {"ok":True}

@router.post("/cuentas-transferencia/{cuenta_id}/activar")
async def activar_cuenta(cuenta_id: int, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    # Deactivate all first
    await db.execute(text("UPDATE cuentas_transferencia SET activa = false"))
    # Activate selected
    await db.execute(text("UPDATE cuentas_transferencia SET activa = true WHERE id = :id"), {"id": cuenta_id})
    await db.commit()
    return {"ok":True}

@router.delete("/cuentas-transferencia/{cuenta_id}")
async def eliminar_cuenta(cuenta_id: int, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    result = await db.execute(select(CuentaTransferencia).where(CuentaTransferencia.id == cuenta_id))
    c = result.scalar_one_or_none()
    if not c: raise HTTPException(status_code=404, detail="No encontrada")
    if c.activa: raise HTTPException(status_code=400, detail="No se puede eliminar la cuenta activa")
    await db.delete(c)
    await db.commit()
    return {"ok":True}


# ══════ CORTE DE CAJA PDF ══════

@router.post("/finanzas/corte-pdf")
async def corte_pdf(
    desde: str, hasta: str,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    d_desde = date.fromisoformat(desde)
    d_hasta = date.fromisoformat(hasta)
    verde = colors.HexColor('#193a2c')
    dorado = colors.HexColor('#d4a843')
    gris = colors.HexColor('#f0ede8')

    # Fetch ingresos
    ri = await db.execute(text(f"""
        SELECT p.numero, p.pago_confirmado_at::date, COALESCE(c.nombre,'Mostrador') as cli, p.canal, p.forma_pago, p.total
        FROM pedidos p LEFT JOIN clientes c ON c.id=p.customer_id
        WHERE p.pago_confirmado_at::date BETWEEN :d AND :h AND {_VENTAS_WHERE}
        ORDER BY p.pago_confirmado_at, p.id
    """), _dp(desde, hasta))
    ingresos = ri.fetchall()
    total_ing = sum(r[5] for r in ingresos)

    # Otros ingresos
    ro = await db.execute(text("SELECT fecha, concepto, monto, metodo_pago FROM otros_ingresos WHERE fecha BETWEEN :d AND :h ORDER BY fecha"), _dp(desde, hasta))
    otros = ro.fetchall()
    total_otros = sum(r[2] for r in otros)

    # Fetch egresos
    re = await db.execute(text("SELECT fecha, concepto, categoria, proveedor, metodo_pago, monto FROM egresos WHERE fecha BETWEEN :d AND :h ORDER BY fecha"), _dp(desde, hasta))
    egresos_rows = re.fetchall()
    total_egr = sum(r[5] for r in egresos_rows)

    # Build PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], textColor=verde, fontSize=18, spaceAfter=6)
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], textColor=verde, fontSize=13, spaceAfter=8, spaceBefore=16)
    normal = styles['Normal']
    small = ParagraphStyle('Small', parent=normal, fontSize=8, textColor=colors.gray)

    elements = []
    # Header
    elements.append(Paragraph('Florería Lucy', title_style))
    elements.append(Paragraph(f'Corte de caja — {desde} al {hasta}', ParagraphStyle('Sub', parent=normal, fontSize=11, textColor=colors.gray)))
    elements.append(Spacer(1, 0.4*cm))

    def fmt(centavos):
        return f'${centavos/100:,.2f}'

    # INGRESOS
    elements.append(Paragraph('INGRESOS (Ventas)', h2_style))
    if ingresos:
        data_ing = [['Folio', 'Fecha', 'Cliente', 'Canal', 'Pago', 'Total']]
        for r in ingresos:
            data_ing.append([r[0], str(r[1]), r[2][:25], r[3] or '', (r[4] or '')[:15], fmt(r[5])])
        data_ing.append(['', '', '', '', 'SUBTOTAL', fmt(total_ing)])
        t = Table(data_ing, colWidths=[1.3*inch, 0.8*inch, 1.5*inch, 0.7*inch, 1*inch, 0.9*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), verde), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 8), ('FONTSIZE', (0,0), (-1,0), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('BACKGROUND', (0,-1), (-1,-1), gris), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph('Sin ventas en este período', normal))

    if otros:
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph('Otros ingresos', h2_style))
        data_oi = [['Fecha', 'Concepto', 'Pago', 'Monto']]
        for r in otros:
            data_oi.append([str(r[0]), r[1][:30], (r[3] or '')[:15], fmt(r[2])])
        data_oi.append(['', '', 'SUBTOTAL', fmt(total_otros)])
        t2 = Table(data_oi, colWidths=[1*inch, 2.5*inch, 1.2*inch, 1*inch])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), dorado), ('TEXTCOLOR', (0,0), (-1,0), verde),
            ('FONTSIZE', (0,0), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('BACKGROUND', (0,-1), (-1,-1), gris), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
        ]))
        elements.append(t2)

    # EGRESOS
    elements.append(Paragraph('EGRESOS', h2_style))
    if egresos_rows:
        data_eg = [['Fecha', 'Concepto', 'Categoría', 'Proveedor', 'Pago', 'Monto']]
        for r in egresos_rows:
            data_eg.append([str(r[0]), r[1][:25], (r[2] or '')[:15], (r[3] or '')[:15], (r[4] or '')[:15], fmt(r[5])])
        data_eg.append(['', '', '', '', 'SUBTOTAL', fmt(total_egr)])
        t3 = Table(data_eg, colWidths=[0.8*inch, 1.3*inch, 1*inch, 1*inch, 1*inch, 0.9*inch])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ef4444')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('BACKGROUND', (0,-1), (-1,-1), gris), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
        ]))
        elements.append(t3)
    else:
        elements.append(Paragraph('Sin egresos en este período', normal))

    # FLUJO DE CAJA
    elements.append(Paragraph('FLUJO DE CAJA', h2_style))
    total_ing_all = total_ing + total_otros
    saldo = total_ing_all - total_egr
    saldo_color = verde if saldo >= 0 else colors.HexColor('#ef4444')
    data_fc = [
        ['Total ingresos (ventas + otros)', fmt(total_ing_all)],
        ['Total egresos', fmt(total_egr)],
        ['SALDO DEL PERÍODO', fmt(saldo)],
    ]
    t4 = Table(data_fc, colWidths=[4*inch, 1.5*inch])
    t4.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 10), ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('FONTSIZE', (0,-1), (-1,-1), 12),
        ('TEXTCOLOR', (-1,-1), (-1,-1), saldo_color),
        ('LINEBELOW', (0,-2), (-1,-2), 1, colors.lightgrey),
        ('LINEABOVE', (0,-1), (-1,-1), 2, verde),
    ]))
    elements.append(t4)

    # Ingresos por método de pago (cómo pagaron los clientes)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph('Ingresos por método de pago', ParagraphStyle('H3', parent=normal, fontSize=10, textColor=verde, fontName='Helvetica-Bold', spaceBefore=8)))
    mp_ing = {}
    for r in ingresos:
        mp = r[4] or 'Sin info'
        mp_ing[mp] = mp_ing.get(mp, 0) + r[5]
    for r in otros:
        mp = r[3] or 'Sin info'
        mp_ing[mp] = mp_ing.get(mp, 0) + r[2]
    if mp_ing:
        data_mp_ing = [['Método de pago', 'Total recibido']]
        for mp in sorted(mp_ing, key=lambda x: mp_ing[x], reverse=True):
            data_mp_ing.append([mp, fmt(mp_ing[mp])])
        data_mp_ing.append(['TOTAL INGRESOS', fmt(total_ing + total_otros)])
        t5 = Table(data_mp_ing, colWidths=[3*inch, 2*inch])
        t5.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), verde), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
            ('BACKGROUND', (0,-1), (-1,-1), gris), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        elements.append(t5)

    # Egresos por fuente de pago (de dónde sale el dinero)
    mp_egr = {}
    for r in egresos_rows:
        mp = r[4] or 'Sin info'
        mp_egr[mp] = mp_egr.get(mp, 0) + r[5]

    # Desglose de ingresos por canal
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph('Ingresos por canal', ParagraphStyle('H3b', parent=normal, fontSize=10, textColor=verde, fontName='Helvetica-Bold', spaceBefore=8)))
    canal_map = {}
    for r in ingresos:
        c = r[3] or 'Sin info'
        canal_map[c] = canal_map.get(c, 0) + r[5]
    if canal_map:
        data_canal = [['Canal', 'Total', '% del total']]
        for c in sorted(canal_map, key=lambda x: canal_map[x], reverse=True):
            pct = (canal_map[c] / total_ing * 100) if total_ing else 0
            data_canal.append([c, fmt(canal_map[c]), f'{pct:.1f}%'])
        t_canal = Table(data_canal, colWidths=[2.5*inch, 1.5*inch, 1*inch])
        t_canal.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), verde), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ]))
        elements.append(t_canal)

    # Egresos por fuente (de dónde salió el dinero)
    if mp_egr:
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph('Egresos por fuente de pago', ParagraphStyle('H3c', parent=normal, fontSize=10, textColor=colors.HexColor('#ef4444'), fontName='Helvetica-Bold', spaceBefore=8)))
        data_egr_mp = [['Fuente', 'Total gastado']]
        for mp in sorted(mp_egr, key=lambda x: mp_egr[x], reverse=True):
            data_egr_mp.append([mp, fmt(mp_egr[mp])])
        data_egr_mp.append(['TOTAL EGRESOS', fmt(total_egr)])
        t_egr_mp = Table(data_egr_mp, colWidths=[3*inch, 2*inch])
        t_egr_mp.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ef4444')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
            ('BACKGROUND', (0,-1), (-1,-1), gris), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        elements.append(t_egr_mp)

    # Footer
    ahora = datetime.now(TZ)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f'Generado: {ahora.strftime("%d/%m/%Y %H:%M")} — Florería Lucy', small))

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=corte_{desde}_{hasta}.pdf'})


# ══════ ESTADÍSTICAS ══════
# Las estadísticas reflejan VENTAS (cuándo entró el dinero), no entregas.
# Filtro por pago_confirmado_at::date (auditado 11-abr-2026).

_VENTAS_WHERE = (
    "estado NOT IN ('Cancelado','rechazado') "
    "AND pago_confirmado = true "
    "AND pago_confirmado_at IS NOT NULL"
)
_FECHA_VENTA = "pago_confirmado_at::date"

def _dp(desde: str, hasta: str):
    """Parse date strings to date objects for asyncpg bind params."""
    return {"d": date.fromisoformat(desde), "h": date.fromisoformat(hasta)}

@router.get("/estadisticas/facturacion")
async def est_facturacion(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    # Total ventas (filtrado por fecha de pago confirmado)
    r = await db.execute(text(f"SELECT COALESCE(SUM(total),0),COUNT(*) FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}"), _dp(desde,hasta))
    total_ventas, num_ventas = r.one()
    # Otros ingresos
    r2 = await db.execute(text("SELECT COALESCE(SUM(monto),0) FROM otros_ingresos WHERE fecha BETWEEN :d AND :h"), _dp(desde,hasta))
    total_otros = r2.scalar() or 0
    total = total_ventas + total_otros
    # vs anterior
    dias = (date.fromisoformat(hasta) - date.fromisoformat(desde)).days + 1
    ant_hasta = (date.fromisoformat(desde) - timedelta(days=1)).isoformat()
    ant_desde = (date.fromisoformat(desde) - timedelta(days=dias)).isoformat()
    r3 = await db.execute(text(f"SELECT COALESCE(SUM(total),0) FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}"), _dp(ant_desde,ant_hasta))
    r3b = await db.execute(text("SELECT COALESCE(SUM(monto),0) FROM otros_ingresos WHERE fecha BETWEEN :d AND :h"), _dp(ant_desde,ant_hasta))
    total_ant = (r3.scalar() or 0) + (r3b.scalar() or 0)
    vs = round((total - total_ant) / total_ant * 100, 1) if total_ant else 0
    # Por día (fecha de venta)
    rd = await db.execute(text(f"SELECT {_FECHA_VENTA} as f, COALESCE(SUM(total),0) as t, COUNT(*) as c FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE} GROUP BY f ORDER BY f"), _dp(desde,hasta))
    por_dia = [{"fecha":str(r[0]),"total":r[1],"ventas":r[2]} for r in rd.fetchall()]
    # Por hora (consistente: hora del pago confirmado)
    rh = await db.execute(text(f"SELECT EXTRACT(HOUR FROM pago_confirmado_at) as h, COALESCE(SUM(total),0) as t, COUNT(*) as c FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE} GROUP BY h ORDER BY h"), _dp(desde,hasta))
    por_hora = [{"hora":int(r[0]),"total":r[1],"ventas":r[2]} for r in rh.fetchall()]
    return {"total":total,"total_ventas":total_ventas,"total_otros":total_otros,"num_ventas":num_ventas,"vs_anterior":vs,"por_dia":por_dia,"por_hora":por_hora}

@router.get("/estadisticas/ticket-medio")
async def est_ticket(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    r = await db.execute(text(f"SELECT COALESCE(SUM(total),0),COUNT(*) FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}"), _dp(desde,hasta))
    total, count = r.one()
    valor = round(total / count) if count else 0
    # vs anterior
    dias = (date.fromisoformat(hasta) - date.fromisoformat(desde)).days + 1
    ad = (date.fromisoformat(desde) - timedelta(days=dias)).isoformat()
    ah = (date.fromisoformat(desde) - timedelta(days=1)).isoformat()
    ra = await db.execute(text(f"SELECT COALESCE(SUM(total),0),COUNT(*) FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}"), _dp(ad,ah))
    ta, ca = ra.one()
    va = round(ta/ca) if ca else 0
    vs = round((valor-va)/va*100,1) if va else 0
    rd = await db.execute(text(f"SELECT {_FECHA_VENTA} as f, COALESCE(AVG(total),0)::int as avg, MIN(total) as mn, MAX(total) as mx FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE} GROUP BY f ORDER BY f"), _dp(desde,hasta))
    por_dia = [{"fecha":str(r[0]),"promedio":r[1],"min":r[2],"max":r[3]} for r in rd.fetchall()]
    return {"valor":valor,"vs_anterior":vs,"por_dia":por_dia}

@router.get("/estadisticas/ganancia")
async def est_ganancia(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    r = await db.execute(text(f"""
        SELECT {_FECHA_VENTA} as f, ped.total as venta,
               COALESCE(SUM(ip.cantidad * COALESCE(p.costo_unitario*100, p.costo, 0)),0)::bigint as costo_total,
               COUNT(CASE WHEN p.costo_unitario IS NULL AND p.costo=0 THEN 1 END) as sin_costo
        FROM pedidos ped JOIN items_pedido ip ON ip.pedido_id=ped.id
        JOIN productos p ON p.id=ip.producto_id
        WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}
        GROUP BY ped.id, ped.pago_confirmado_at, ped.total
    """), _dp(desde,hasta))
    rows = r.fetchall()
    total_venta = sum(r[1] for r in rows)
    total_costo = sum(r[2] for r in rows)
    total_ganancia = total_venta - total_costo
    sin_costo = sum(1 for r in rows if r[3] > 0)
    # Por día
    dia_map = {}
    for r in rows:
        d = str(r[0])
        if d not in dia_map: dia_map[d] = {"fecha":d,"facturacion":0,"costo":0,"ganancia":0}
        dia_map[d]["facturacion"] += r[1]; dia_map[d]["costo"] += r[2]
        dia_map[d]["ganancia"] = dia_map[d]["facturacion"] - dia_map[d]["costo"]
    por_dia = sorted(dia_map.values(), key=lambda x: x["fecha"])
    for d in por_dia: d["margen"] = round(d["ganancia"]/d["facturacion"]*100,1) if d["facturacion"] else 0
    return {"total":total_ganancia,"facturacion":total_venta,"costo":total_costo,"productos_sin_costo":sin_costo,"por_dia":por_dia}

@router.get("/estadisticas/medios-pago")
async def est_medios(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    r = await db.execute(text(f"SELECT COALESCE(forma_pago,'Sin info') as mp, COUNT(*) as c, COALESCE(SUM(total),0) as t FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE} AND forma_pago IS NOT NULL AND forma_pago != '' AND forma_pago != '0' GROUP BY forma_pago ORDER BY t DESC"), _dp(desde,hasta))
    rows = [{"metodo":r[0],"count":r[1],"total":r[2]} for r in r.fetchall()]
    grand = sum(r["total"] for r in rows) or 1
    for r in rows: r["porcentaje"] = round(r["total"]/grand*100,1)
    mas_usado = rows[0]["metodo"] if rows else "—"
    return {"mas_usado":mas_usado,"distribucion":rows}

@router.get("/estadisticas/productos-top")
async def est_productos(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    r = await db.execute(text(f"""
        SELECT p.nombre, p.categoria, SUM(ip.cantidad) as qty, SUM(ip.cantidad*ip.precio_unitario) as val
        FROM items_pedido ip JOIN pedidos ped ON ped.id=ip.pedido_id JOIN productos p ON p.id=ip.producto_id
        WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}
        GROUP BY p.nombre, p.categoria ORDER BY val DESC LIMIT 10
    """), _dp(desde,hasta))
    por_valor = [{"nombre":r[0],"categoria":r[1],"cantidad":int(r[2]),"total":r[3]} for r in r.fetchall()]
    r2 = await db.execute(text(f"""
        SELECT p.nombre, p.categoria, SUM(ip.cantidad) as qty, SUM(ip.cantidad*ip.precio_unitario) as val
        FROM items_pedido ip JOIN pedidos ped ON ped.id=ip.pedido_id JOIN productos p ON p.id=ip.producto_id
        WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}
        GROUP BY p.nombre, p.categoria ORDER BY qty DESC LIMIT 10
    """), _dp(desde,hasta))
    por_cantidad = [{"nombre":r[0],"categoria":r[1],"cantidad":int(r[2]),"total":r[3]} for r in r2.fetchall()]
    return {"por_valor":por_valor,"por_cantidad":por_cantidad}

@router.get("/estadisticas/clientes-top")
async def est_clientes(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    r = await db.execute(text(f"""
        SELECT c.nombre, COUNT(ped.id) as n, SUM(ped.total) as t, MAX(ped.pago_confirmado_at) as ult
        FROM pedidos ped JOIN clientes c ON c.id=ped.customer_id
        WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}
        GROUP BY c.id, c.nombre ORDER BY t DESC LIMIT 10
    """), _dp(desde,hasta))
    por_valor = [{"nombre":r[0],"pedidos":r[1],"total":r[2],"ticket_medio":round(r[2]/r[1]) if r[1] else 0,"ultima":str(r[3]) if r[3] else None} for r in r.fetchall()]
    r2 = await db.execute(text(f"""
        SELECT c.nombre, COUNT(ped.id) as n, SUM(ped.total) as t, MAX(ped.pago_confirmado_at) as ult
        FROM pedidos ped JOIN clientes c ON c.id=ped.customer_id
        WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE}
        GROUP BY c.id, c.nombre ORDER BY n DESC LIMIT 10
    """), _dp(desde,hasta))
    por_compras = [{"nombre":r[0],"pedidos":r[1],"total":r[2],"ticket_medio":round(r[2]/r[1]) if r[1] else 0,"ultima":str(r[3]) if r[3] else None} for r in r2.fetchall()]
    return {"por_valor":por_valor,"por_compras":por_compras}

@router.get("/estadisticas/canales")
async def est_canales(desde: str, hasta: str, panel_session: str | None = Cookie(default=None), db: AsyncSession = Depends(get_db)):
    _auth(panel_session)
    r = await db.execute(text(f"SELECT canal, COUNT(*) as c, COALESCE(SUM(total),0) as t FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE} GROUP BY canal"), _dp(desde,hasta))
    rows = [{"canal":r[0],"count":r[1],"total":r[2]} for r in r.fetchall()]
    grand = sum(r["total"] for r in rows) or 1
    for r in rows: r["porcentaje"] = round(r["total"]/grand*100,1)
    rd = await db.execute(text(f"SELECT {_FECHA_VENTA} as f, canal, COUNT(*) as c FROM pedidos WHERE {_FECHA_VENTA} BETWEEN :d AND :h AND {_VENTAS_WHERE} GROUP BY f, canal ORDER BY f"), _dp(desde,hasta))
    por_dia = [{"fecha":str(r[0]),"canal":r[1],"count":r[2]} for r in rd.fetchall()]
    return {"distribucion":rows,"por_dia":por_dia}


# ══════ SUBIDA IMÁGENES ══════

@router.post("/productos/imagen")
async def subir_imagen_producto(
    imagen: UploadFile = File(...),
    panel_session: str | None = Cookie(default=None),
):
    _auth(panel_session)
    contents = await imagen.read()
    result = cloudinary.uploader.upload(
        contents,
        folder="productos",
        public_id=f"prod_{datetime.now(TZ).strftime('%Y%m%d%H%M%S')}",
    )
    return {"url": result["secure_url"]}


# ══════ EXPORTAR / IMPORTAR PRODUCTOS ══════

@router.get("/productos/exportar")
async def exportar_productos(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    from fastapi.responses import StreamingResponse
    import io, openpyxl

    result = await db.execute(select(Producto).order_by(Producto.categoria, Producto.nombre))
    productos = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    headers = ["Nombre", "Categoría", "Código", "Descripción", "Precio",
               "Precio de promoción", "Mostrar en el catálogo", "Controlar stock", "Stock actual",
               "Costo unitario", "Alto (cm)", "Ancho (cm)"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    for p in productos:
        ws.append([
            p.nombre,
            p.categoria,
            p.codigo or "",
            p.descripcion or "",
            round(p.precio / 100, 2) if p.precio else 0,
            round(p.precio_descuento / 100, 2) if p.precio_descuento else "",
            "S" if p.visible_catalogo else "N",
            "S" if p.stock_activo else "N",
            p.stock if p.stock_activo else "",
            float(p.costo_unitario) if p.costo_unitario else "",
            float(p.medida_alto) if p.medida_alto else "",
            float(p.medida_ancho) if p.medida_ancho else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    hoy = datetime.now(TZ).strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=productos_floreria_lucy_{hoy}.xlsx"},
    )


@router.post("/productos/importar")
async def importar_productos(
    archivo: UploadFile = File(...),
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    import openpyxl, io

    contents = await archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    actualizados = 0
    creados = 0
    errores = 0

    for row in rows:
        try:
            if len(row) < 5:
                errores += 1
                continue
            nombre = str(row[0] or "").strip()
            categoria = str(row[1] or "").strip()
            codigo = str(row[2] or "").strip()
            descripcion = str(row[3] or "").strip() or None
            precio = int(round(float(row[4] or 0) * 100))
            precio_desc = int(round(float(row[5]) * 100)) if row[5] not in (None, "", "N/A") else None
            visible = str(row[6] or "S").strip().upper() == "S" if len(row) > 6 else True
            stock_activo = str(row[7] or "N").strip().upper() == "S" if len(row) > 7 else False
            stock_val = int(row[8] or 0) if len(row) > 8 and row[8] not in (None, "") else 0
            costo_u = float(row[9]) if len(row) > 9 and row[9] not in (None, "") else None
            m_alto = float(row[10]) if len(row) > 10 and row[10] not in (None, "") else None
            m_ancho = float(row[11]) if len(row) > 11 and row[11] not in (None, "") else None

            if not nombre or not codigo:
                errores += 1
                continue

            # Ensure category exists
            cat_result = await db.execute(select(Categoria).where(Categoria.nombre == categoria))
            if not cat_result.scalar_one_or_none():
                db.add(Categoria(nombre=categoria, tipo="normal", orden=0))
                await db.flush()

            # Find by codigo
            result = await db.execute(select(Producto).where(Producto.codigo == codigo))
            prod = result.scalar_one_or_none()

            if prod:
                prod.nombre = nombre
                prod.categoria = categoria
                prod.descripcion = descripcion
                prod.precio = precio
                prod.precio_descuento = precio_desc
                prod.visible_catalogo = visible
                prod.stock_activo = stock_activo
                prod.stock = stock_val
                prod.costo_unitario = costo_u
                prod.medida_alto = m_alto
                prod.medida_ancho = m_ancho
                actualizados += 1
            else:
                prod = Producto(
                    nombre=nombre, categoria=categoria, codigo=codigo,
                    descripcion=descripcion, precio=precio, precio_descuento=precio_desc,
                    visible_catalogo=visible, stock_activo=stock_activo, stock=stock_val,
                    costo_unitario=costo_u, medida_alto=m_alto, medida_ancho=m_ancho,
                    activo=True, disponible_hoy=True,
                )
                db.add(prod)
                creados += 1
        except Exception:
            errores += 1

    await db.commit()
    return {"ok": True, "actualizados": actualizados, "creados": creados, "errores": errores}


# ══════ CATEGORÍAS ══════

@router.get("/categorias")
async def listar_categorias(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(Categoria).order_by(Categoria.orden, Categoria.nombre))
    cats = result.scalars().all()
    items = []
    for c in cats:
        count = await db.execute(
            select(func.count(Producto.id)).where(Producto.categoria == c.nombre, Producto.activo == True)
        )
        items.append({
            "id": c.id, "nombre": c.nombre, "tipo": c.tipo,
            "orden": c.orden, "productos_activos": count.scalar() or 0,
        })
    return items


@router.post("/categorias")
async def crear_categoria(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    c = Categoria(
        nombre=data["nombre"].strip(),
        tipo=data.get("tipo", "normal"),
        orden=data.get("orden", 0),
    )
    db.add(c)
    try:
        await db.commit()
        await db.refresh(c)
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=400, detail="Categoría ya existe")
    return {"ok": True, "id": c.id}


@router.put("/categorias/{cat_id}")
async def actualizar_categoria(
    cat_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(Categoria).where(Categoria.id == cat_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    data = await request.json()
    old_nombre = c.nombre
    for k in ["nombre", "tipo", "orden"]:
        if k in data:
            setattr(c, k, data[k])
    # Update products if category name changed
    if "nombre" in data and data["nombre"] != old_nombre:
        await db.execute(
            text("UPDATE productos SET categoria = :new WHERE categoria = :old"),
            {"new": data["nombre"], "old": old_nombre},
        )
    await db.commit()
    return {"ok": True}


@router.delete("/categorias/{cat_id}")
async def eliminar_categoria(
    cat_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(Categoria).where(Categoria.id == cat_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    count = await db.execute(
        select(func.count(Producto.id)).where(Producto.categoria == c.nombre, Producto.activo == True)
    )
    if (count.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="No se puede eliminar: tiene productos activos")
    await db.delete(c)
    await db.commit()
    return {"ok": True}


# ══════ VARIANTES ══════

@router.get("/variantes/{producto_id}")
async def listar_variantes(
    producto_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(
        select(ProductoVariante)
        .where(ProductoVariante.producto_id == producto_id)
        .order_by(ProductoVariante.tipo, ProductoVariante.nombre)
    )
    return [
        {"id": v.id, "producto_id": v.producto_id, "tipo": v.tipo, "nombre": v.nombre,
         "codigo": v.codigo, "imagen_url": v.imagen_url, "precio": v.precio,
         "precio_descuento": v.precio_descuento, "stock_activo": v.stock_activo,
         "stock": v.stock, "activo": v.activo}
        for v in result.scalars().all()
    ]


@router.post("/variantes/{producto_id}")
async def crear_variante(
    producto_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    v = ProductoVariante(
        producto_id=producto_id,
        tipo=data["tipo"],
        nombre=data["nombre"],
        codigo=data.get("codigo"),
        imagen_url=data.get("imagen_url"),
        precio=data.get("precio", 0),
        precio_descuento=data.get("precio_descuento"),
        stock_activo=data.get("stock_activo", False),
        stock=data.get("stock", 0),
        activo=data.get("activo", True),
    )
    db.add(v)
    await db.commit()
    await db.refresh(v)
    return {"ok": True, "id": v.id}


@router.put("/variantes/{variante_id}")
async def actualizar_variante(
    variante_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(ProductoVariante).where(ProductoVariante.id == variante_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Variante no encontrada")
    data = await request.json()
    for k in ["tipo", "nombre", "codigo", "imagen_url", "precio", "precio_descuento", "stock_activo", "stock", "activo"]:
        if k in data:
            setattr(v, k, data[k])
    await db.commit()
    return {"ok": True}


@router.delete("/variantes/{variante_id}")
async def eliminar_variante(
    variante_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(select(ProductoVariante).where(ProductoVariante.id == variante_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Variante no encontrada")
    await db.delete(v)
    await db.commit()
    return {"ok": True}


# ══════ GENERAR DESCRIPCIÓN IA ══════

@router.post("/productos/generar-descripcion")
async def generar_descripcion(
    request: Request,
    panel_session: str | None = Cookie(default=None),
):
    _auth(panel_session)
    import httpx
    data = await request.json()
    nombre = data.get("nombre", "")
    categoria = data.get("categoria", "")
    desc_base = data.get("descripcion_base", "")

    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="API key de Anthropic no configurada")

    if desc_base:
        system = "Mejora y expande esta descripción de producto para una florería boutique. Mantén el tono elegante. Máximo 3 oraciones. Responde solo con la descripción mejorada."
        user_msg = f"Producto: {nombre} ({categoria})\nDescripción actual: {desc_base}"
    else:
        system = "Eres el asistente de Florería Lucy, una florería boutique en Chihuahua, México. Genera una descripción de producto corta, elegante y atractiva (máximo 3 oraciones) para uso en catálogo online. Responde solo con la descripción, sin comillas ni prefijos."
        user_msg = f"Producto: {nombre}\nCategoría: {categoria}"

    async with httpx.AsyncClient() as client:
        r = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 300,
                "system": system,
                "messages": [{"role": "user", "content": user_msg}],
            },
            timeout=30,
        )
        d = r.json()
        descripcion = d["content"][0]["text"] if d.get("content") else ""
        return {"descripcion": descripcion}


# ══════ CAMBIAR ESTADO PEDIDO (admin) ══════

@router.patch("/pedido/{pedido_id}/estado")
async def admin_cambiar_estado(
    pedido_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """Cambiar estado de un pedido desde admin. No envía WhatsApp."""
    _auth(panel_session)
    data = await request.json()
    nuevo_estado = data.get("estado")
    if not nuevo_estado:
        raise HTTPException(status_code=400, detail="Falta el campo 'estado'")

    from app.core.estados import EstadoPedido as EP
    estados_validos = [
        EP.ESPERANDO_VALIDACION, EP.PENDIENTE_PAGO, EP.COMPROBANTE_RECIBIDO,
        EP.PAGADO, EP.EN_PRODUCCION, EP.LISTO, EP.LISTO_TALLER,
        EP.EN_CAMINO, EP.ENTREGADO, EP.CANCELADO, EP.INTENTO_FALLIDO,
    ]
    if nuevo_estado not in estados_validos:
        raise HTTPException(status_code=400, detail=f"Estado no válido: {nuevo_estado}")

    result = await db.execute(select(Pedido).where(Pedido.id == pedido_id))
    pedido = result.scalar_one_or_none()
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")

    estado_anterior = pedido.estado
    pedido.estado = nuevo_estado

    # Si se marca como entregado, registrar timestamp
    if nuevo_estado == EP.ENTREGADO:
        from app.core.utils import ahora
        pedido.entregado_at = ahora()

    await db.commit()
    logger.info(f"Admin cambió estado pedido {pedido.numero}: {estado_anterior} → {nuevo_estado}")
    return {"ok": True, "folio": pedido.numero, "estado_anterior": estado_anterior, "estado_nuevo": nuevo_estado}


# ══════ HISTORIAL DE STOCK ══════

@router.get("/stock-historial")
async def stock_historial(
    categoria: str | None = None,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    filtro_cat = "AND p.categoria = :cat" if categoria else ""
    params = {"cat": categoria} if categoria else {}
    result = await db.execute(text(f"""
        SELECT p.id, p.nombre, p.categoria, p.imagen_url, p.stock,
               COALESCE(SUM(ip.cantidad), 0) as vendidos
        FROM productos p
        LEFT JOIN items_pedido ip ON ip.producto_id = p.id
        LEFT JOIN pedidos ped ON ped.id = ip.pedido_id
          AND {_VENTAS_WHERE}
        WHERE p.stock_activo = true {filtro_cat}
        GROUP BY p.id, p.nombre, p.categoria, p.imagen_url, p.stock
        ORDER BY p.categoria, p.nombre
    """), params)
    rows = result.fetchall()
    return [
        {
            "id": r[0],
            "nombre": r[1],
            "categoria": r[2],
            "imagen_url": r[3],
            "stock_actual": r[4],
            "vendidos": int(r[5]),
            "stock_inicial_estimado": r[4] + int(r[5]),
        }
        for r in rows
    ]


# ══════ CUENTAS FINANCIERAS (Caja + Caja Chica) ══════

@router.get("/cuentas-financieras")
async def listar_cuentas_financieras(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    result = await db.execute(text(
        "SELECT id, nombre, tipo, saldo_inicial, fecha_inicio, fondo_base, activo "
        "FROM cuentas_financieras ORDER BY id"
    ))
    return [
        {"id": r[0], "nombre": r[1], "tipo": r[2], "saldo_inicial": r[3],
         "fecha_inicio": str(r[4]), "fondo_base": r[5], "activo": r[6]}
        for r in result.fetchall()
    ]


@router.post("/cuentas-financieras")
async def crear_cuenta_financiera(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    try:
        fecha_ini = data.get("fecha_inicio") or datetime.now(TZ).date().isoformat()
        if isinstance(fecha_ini, str):
            fecha_ini = date.fromisoformat(fecha_ini)
        await db.execute(text("""
            INSERT INTO cuentas_financieras (nombre, tipo, saldo_inicial, fecha_inicio, fondo_base, activo)
            VALUES (:n, :t, :s, :f, :fb, true)
            ON CONFLICT (nombre) DO NOTHING
        """), {
            "n": data["nombre"], "t": data["tipo"],
            "s": int(data.get("saldo_inicial", 0)),
            "f": fecha_ini,
            "fb": int(data.get("fondo_base", 0)),
        })
        await db.commit()
        return {"ok": True}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cuentas-financieras/seed-default")
async def seed_default_cuentas(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """Crea tablas + seed Caja y Caja Chica si no existen. Idempotente.
    Usa SQL crudo para no depender del create_all de SQLAlchemy."""
    _auth(panel_session)
    try:
        # 1. Crear tablas si no existen
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS cuentas_financieras (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(100) UNIQUE NOT NULL,
                tipo VARCHAR(20) NOT NULL,
                saldo_inicial INTEGER DEFAULT 0,
                fecha_inicio DATE NOT NULL,
                fondo_base INTEGER DEFAULT 0,
                activo BOOLEAN DEFAULT true,
                created_at TIMESTAMP
            )
        """))
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS movimientos_cuenta (
                id SERIAL PRIMARY KEY,
                cuenta_id INTEGER NOT NULL,
                fecha DATE NOT NULL,
                tipo VARCHAR(30) NOT NULL,
                concepto VARCHAR(200) NOT NULL,
                monto INTEGER NOT NULL,
                cuenta_destino_id INTEGER,
                referencia_tipo VARCHAR(50),
                referencia_id INTEGER,
                notas TEXT,
                created_at TIMESTAMP
            )
        """))
        # 2. Asegurar columna cuenta_id en egresos
        await db.execute(text("ALTER TABLE egresos ADD COLUMN IF NOT EXISTS cuenta_id INTEGER"))
        # 3. Seed cuentas
        hoy = datetime.now(TZ).date()
        seeds = [
            ("Caja", "caja", 0, 100000),
            ("Caja Chica", "caja_chica", 0, 0),
        ]
        creadas = 0
        for nombre, tipo, saldo, fondo in seeds:
            exists = (await db.execute(text(
                "SELECT id FROM cuentas_financieras WHERE nombre = :n"
            ), {"n": nombre})).fetchone()
            if not exists:
                await db.execute(text("""
                    INSERT INTO cuentas_financieras (nombre, tipo, saldo_inicial, fecha_inicio, fondo_base, activo, created_at)
                    VALUES (:n, :t, :s, :f, :fb, true, NOW())
                """), {"n": nombre, "t": tipo, "s": saldo, "f": hoy, "fb": fondo})
                creadas += 1
        await db.commit()
        return {"ok": True, "creadas": creadas}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Seed cuentas: {str(e)}")


@router.put("/cuentas-financieras/{cuenta_id}")
async def actualizar_cuenta_financiera(
    cuenta_id: int,
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    sets = []
    params = {"id": cuenta_id}
    for k in ["nombre", "saldo_inicial", "fondo_base", "activo"]:
        if k in data:
            sets.append(f"{k} = :{k}")
            params[k] = data[k]
    if "fecha_inicio" in data:
        sets.append("fecha_inicio = :fecha_inicio")
        params["fecha_inicio"] = (
            date.fromisoformat(data["fecha_inicio"])
            if isinstance(data["fecha_inicio"], str) else data["fecha_inicio"]
        )
    if not sets:
        return {"ok": True}
    await db.execute(text(
        f"UPDATE cuentas_financieras SET {', '.join(sets)} WHERE id = :id"
    ), params)
    await db.commit()
    return {"ok": True}


async def _saldo_cuenta(db: AsyncSession, cuenta: tuple) -> dict:
    """Calcula el saldo actual de una cuenta financiera.

    Para tipo='caja': suma automáticamente los pedidos pagados en efectivo
    desde fecha_inicio (no requiere registrar movimientos manuales).
    Para cualquier tipo: suma movimientos manuales y resta egresos asignados.
    """
    cid, nombre, tipo, saldo_ini, f_ini, fondo = cuenta
    dep = (await db.execute(text(
        "SELECT COALESCE(SUM(monto),0) FROM movimientos_cuenta "
        "WHERE cuenta_id=:cid AND fecha >= :fi AND tipo IN "
        "('deposito_corte_pos','deposito_manual','transferencia_in','ajuste_positivo')"
    ), {"cid": cid, "fi": f_ini})).scalar() or 0
    ret = (await db.execute(text(
        "SELECT COALESCE(SUM(monto),0) FROM movimientos_cuenta "
        "WHERE cuenta_id=:cid AND fecha >= :fi AND tipo IN "
        "('retiro_manual','transferencia_out','ajuste_negativo')"
    ), {"cid": cid, "fi": f_ini})).scalar() or 0
    # Match egresos por cuenta_id explícito O por método_pago = nombre cuenta
    egr = (await db.execute(text(
        "SELECT COALESCE(SUM(monto),0) FROM egresos "
        "WHERE fecha >= :fi AND ("
        "  cuenta_id = :cid "
        "  OR LOWER(TRIM(COALESCE(metodo_pago,''))) = LOWER(TRIM(:nombre))"
        ")"
    ), {"cid": cid, "fi": f_ini, "nombre": nombre})).scalar() or 0
    # Match otros_ingresos por método_pago = nombre cuenta (no tiene cuenta_id)
    otros_ing = (await db.execute(text(
        "SELECT COALESCE(SUM(monto),0) FROM otros_ingresos "
        "WHERE fecha >= :fi "
        "AND LOWER(TRIM(COALESCE(metodo_pago,''))) = LOWER(TRIM(:nombre))"
    ), {"fi": f_ini, "nombre": nombre})).scalar() or 0
    pos_efectivo = 0
    if tipo == 'caja':
        # Pedidos POS pagados en efectivo desde fecha_inicio
        from datetime import time as time_type
        from app.core.estados import EstadoPedido as _EP
        try:
            estados_venta = list(_EP.VENTA_COMPLETADA)
        except Exception:
            estados_venta = ['pagado', 'En producción', 'listo_taller', 'En camino', 'Entregado']
        f_ini_dt = datetime.combine(f_ini, time_type.min)
        pos_efectivo = (await db.execute(text(
            "SELECT COALESCE(SUM(total),0) FROM pedidos "
            "WHERE pago_confirmado_at >= :fi "
            "AND LOWER(COALESCE(forma_pago,'')) LIKE '%efectivo%' "
            "AND estado = ANY(:estados)"
        ), {"fi": f_ini_dt, "estados": estados_venta})).scalar() or 0
    saldo = (saldo_ini or 0) + int(dep) - int(ret) - int(egr) + int(pos_efectivo) + int(otros_ing)
    return {
        "id": cid, "nombre": nombre, "tipo": tipo,
        "saldo_inicial": saldo_ini, "fecha_inicio": str(f_ini),
        "fondo_base": fondo,
        "depositos": int(dep), "retiros": int(ret),
        "egresos": int(egr), "ingresos_efectivo_pos": int(pos_efectivo),
        "otros_ingresos": int(otros_ing),
        "saldo_actual": saldo,
    }


@router.get("/cuentas-financieras/saldos")
async def saldos_cuentas(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    cuentas = (await db.execute(text(
        "SELECT id, nombre, tipo, saldo_inicial, fecha_inicio, fondo_base "
        "FROM cuentas_financieras WHERE activo = true ORDER BY id"
    ))).fetchall()
    return [await _saldo_cuenta(db, c) for c in cuentas]


@router.get("/movimientos-cuenta")
async def listar_movimientos(
    cuenta_id: int | None = None,
    desde: str | None = None,
    hasta: str | None = None,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    sql = ("SELECT id, cuenta_id, fecha, tipo, concepto, monto, cuenta_destino_id, "
           "referencia_tipo, referencia_id, notas FROM movimientos_cuenta")
    wheres = []
    params = {}
    if cuenta_id:
        wheres.append("(cuenta_id = :cid OR cuenta_destino_id = :cid)")
        params["cid"] = cuenta_id
    if desde:
        wheres.append("fecha >= :desde")
        params["desde"] = date.fromisoformat(desde)
    if hasta:
        wheres.append("fecha <= :hasta")
        params["hasta"] = date.fromisoformat(hasta)
    if wheres:
        sql += " WHERE " + " AND ".join(wheres)
    sql += " ORDER BY fecha DESC, id DESC LIMIT 500"
    result = await db.execute(text(sql), params)
    return [
        {"id": r[0], "cuenta_id": r[1], "fecha": str(r[2]), "tipo": r[3],
         "concepto": r[4], "monto": r[5], "cuenta_destino_id": r[6],
         "referencia_tipo": r[7], "referencia_id": r[8], "notas": r[9]}
        for r in result.fetchall()
    ]


@router.post("/movimientos-cuenta")
async def crear_movimiento(
    request: Request,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    data = await request.json()
    try:
        # Asegurar que la tabla exista (por si el create_all no la creó al arranque)
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS movimientos_cuenta (
                id SERIAL PRIMARY KEY,
                cuenta_id INTEGER NOT NULL,
                fecha DATE NOT NULL,
                tipo VARCHAR(30) NOT NULL,
                concepto VARCHAR(200) NOT NULL,
                monto INTEGER NOT NULL,
                cuenta_destino_id INTEGER,
                referencia_tipo VARCHAR(50),
                referencia_id INTEGER,
                notas TEXT,
                created_at TIMESTAMP
            )
        """))
        fecha_val = (date.fromisoformat(data["fecha"])
                     if isinstance(data["fecha"], str) else data["fecha"])
        await db.execute(text("""
            INSERT INTO movimientos_cuenta
                (cuenta_id, fecha, tipo, concepto, monto, cuenta_destino_id, referencia_tipo, referencia_id, notas, created_at)
            VALUES (:cid, :f, :t, :c, :m, :cd, :rt, :ri, :n, NOW())
        """), {
            "cid": data["cuenta_id"], "f": fecha_val, "t": data["tipo"],
            "c": data.get("concepto", ""), "m": int(data.get("monto", 0)),
            "cd": data.get("cuenta_destino_id"),
            "rt": data.get("referencia_tipo"),
            "ri": data.get("referencia_id"),
            "n": data.get("notas"),
        })
        await db.commit()
        return {"ok": True}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/movimientos-cuenta/{mov_id}")
async def eliminar_movimiento(
    mov_id: int,
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    _auth(panel_session)
    await db.execute(text("DELETE FROM movimientos_cuenta WHERE id = :id"), {"id": mov_id})
    await db.commit()
    return {"ok": True}


@router.post("/cuentas-financieras/cerrar-semana")
async def cerrar_semana_caja(
    panel_session: str | None = Cookie(default=None),
    db: AsyncSession = Depends(get_db),
):
    """Transfiere el saldo de Caja a Caja Chica.
    Crea 2 movimientos (out + in) y resetea fecha_inicio de Caja a mañana."""
    _auth(panel_session)
    try:
        caja_row = (await db.execute(text(
            "SELECT id, nombre, tipo, saldo_inicial, fecha_inicio, fondo_base "
            "FROM cuentas_financieras WHERE tipo='caja' AND activo=true LIMIT 1"
        ))).fetchone()
        chica_row = (await db.execute(text(
            "SELECT id FROM cuentas_financieras WHERE tipo='caja_chica' AND activo=true LIMIT 1"
        ))).fetchone()
        if not caja_row or not chica_row:
            raise HTTPException(status_code=400, detail="Faltan cuentas Caja y/o Caja Chica")
        caja_id = caja_row[0]
        chica_id = chica_row[0]
        saldo_info = await _saldo_cuenta(db, tuple(caja_row))
        saldo_caja = saldo_info["saldo_actual"]
        if saldo_caja <= 0:
            return {"ok": True, "transferido": 0, "mensaje": "Caja en $0, nada que transferir"}
        hoy = datetime.now(TZ).date()
        # Asegurar que la tabla movimientos_cuenta exista
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS movimientos_cuenta (
                id SERIAL PRIMARY KEY,
                cuenta_id INTEGER NOT NULL,
                fecha DATE NOT NULL,
                tipo VARCHAR(30) NOT NULL,
                concepto VARCHAR(200) NOT NULL,
                monto INTEGER NOT NULL,
                cuenta_destino_id INTEGER,
                referencia_tipo VARCHAR(50),
                referencia_id INTEGER,
                notas TEXT,
                created_at TIMESTAMP
            )
        """))
        await db.execute(text("""
            INSERT INTO movimientos_cuenta (cuenta_id, fecha, tipo, concepto, monto, cuenta_destino_id, referencia_tipo, created_at)
            VALUES (:cid, :f, 'transferencia_out', 'Cierre semanal -> Caja Chica', :m, :cd, 'cierre_semanal', NOW())
        """), {"cid": caja_id, "f": hoy, "m": saldo_caja, "cd": chica_id})
        await db.execute(text("""
            INSERT INTO movimientos_cuenta (cuenta_id, fecha, tipo, concepto, monto, cuenta_destino_id, referencia_tipo, created_at)
            VALUES (:cid, :f, 'transferencia_in', 'Cierre semanal <- Caja', :m, :co, 'cierre_semanal', NOW())
        """), {"cid": chica_id, "f": hoy, "m": saldo_caja, "co": caja_id})
        # Reset Caja: fecha_inicio = mañana, saldo_inicial = 0
        manana = hoy + timedelta(days=1)
        await db.execute(text(
            "UPDATE cuentas_financieras SET fecha_inicio = :f, saldo_inicial = 0 WHERE id = :id"
        ), {"f": manana, "id": caja_id})
        await db.commit()
        return {"ok": True, "transferido": saldo_caja}
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Cerrar semana: {type(e).__name__}: {str(e)}")
